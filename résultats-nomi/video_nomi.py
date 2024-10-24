from results_nomi import nb_columns, create_results_sheet, Song
from openpyxl import load_workbook
from pathlib import Path
from moviepy.editor import *
import os
import shutil
from pytube import YouTube
import re
from ffmpeg_normalize import FFmpegNormalize
import json
import configparser


transition = 1
pr = ''
song_per_part = 40
# Set this to True to get max resolution on Youtube videos (slower)
max_resolution = False


config = configparser.ConfigParser()
config.read('../config.txt')
pr_path = config["general"]["pr_path"]


class SampledSong(Song):
    def __init__(self, picker, anime, song_type, info, link, score, sample, sample_length, ranks, order):
        Song.__init__(self, picker, anime, song_type, info, link)
        self.final_score = score
        if sample is None:
            exit("❌ Une ou plusieurs des cases Sample est vide!")
        self.sample = max(sample, transition)
        self.sample_length = sample_length
        if sample_length is None:
            exit("❌ Une ou plusieurs des cases Sample Length est vide!")
        if sample_length <= 0:
            exit("❌ Une ou plusieurs des cases Sample Length contient une valeur invalide ! (<= 0)")
        self.scores = {order[i]: ranks[i] for i in range(len(order))}
        if 'youtu' in link:
            self.extension = 'mp4'
        else:
            self.extension = get_extension(link)
            


def get_result_sheet():
    global pr
    paths = Path('.').glob('**/*.xlsx')
    for path in paths:
        if 'samples' in str(path):
            pr = str(path).split()[0]
            return load_workbook(path)


def get_results(sheet):
    global scoring_pr
    rows = []
    songs = {}
    for row in sheet.iter_rows():
        to_add = []
        for cell in row:
            to_add.append(cell.value)
            if cell.hyperlink:
                to_add.append(cell.hyperlink.target)
        rows.append(to_add)

    header = rows[0]
    scoring_pr = 'Average' in header

    rank_index = header.index('Rank')
    picker_index = header.index('Picker')
    anime_index = header.index('Anime Name')
    type_index = header.index('Song Type')
    info_index = header.index('Song Info')
    score_index = header.index('Score') + 1
    sample_index = header.index('Sample (seconds)') + 1
    sample_length_index = header.index('Sample length (seconds)') + 1
    people_index = max(rank_index, anime_index, type_index, info_index,
                       score_index, sample_index, sample_length_index) + 1

    order = rows[0][people_index-1:]

    for i in range(1, len(rows)):
        if not rows[i][rank_index]:
            raise Exception('Tiebreak needed')
        songs[rows[i][rank_index]] = SampledSong(rows[i][picker_index]  , rows[i][anime_index], rows[i][type_index], rows[i][info_index],
                                                 rows[i][info_index + 1], rows[i][score_index], rows[i][sample_index],
                                                 rows[i][sample_length_index], rows[i][people_index:], order)
    return songs, order


def execute_command(command):
    os.system(command)


def get_extension(link):
    return link.split(".")[-1]


def youtube_dl(link, output_name):
    if (max_resolution):
        yt = YouTube(link)

        # Download highest resolution video
        video_stream = yt.streams.filter(file_extension='mp4', only_video=True).order_by('resolution').desc().first()
        video_path = video_stream.download(filename=f"{output_name}_video.mp4")
        
        # Download highest quality audio
        audio_stream = yt.streams.filter(file_extension='mp4', only_audio=True).order_by('abr').desc().first()
        audio_path = audio_stream.download(filename=f"{output_name}_audio.mp4")
        
        # Merge video and audio
        video_clip = VideoFileClip(video_path)
        audio_clip = AudioFileClip(audio_path)
        
        final_clip = video_clip.set_audio(audio_clip)
        final_clip.write_videofile(filename=f"{output_name}", codec='libx264', audio_codec='aac')
        
        # Clean up temporary files
        video_clip.close()
        audio_clip.close()
        final_clip.close()
        os.remove(video_path)
        os.remove(audio_path)
    else:
        YouTube(link).streams.filter(progressive=True, file_extension='mp4').order_by('resolution').desc().first().download(filename=output_name)


def download_songs(songs, song_range):
    for i in range(song_range[0], song_range[1]):
        if os.path.exists(f"{pr_path}/résultats-nomi/temp/{i}.{songs[i].extension}"):
            print(f"Chanson {i}.{songs[i].extension} trouvée dans le dossier temp, pas besoin de la télécharger !")
        else:
            link = songs[i].link
            print(link)
            if 'catbox' in link or 'animemusicquiz.com' in link:
                command = f'ffmpeg -i {link} -c copy {i}.{songs[i].extension}'
                execute_command(command)
            elif 'youtu' in link:
                tries = 0
                while tries <= 3:
                    try:
                        youtube_dl(link, f'{i}.mp4')
                        tries = 0
                        break
                    except:
                        tries += 1
                        if tries <= 3:
                            print(f"Erreur pendant le téléchargement de {link}, essai numéro {tries}...")
                        else:
                            exit(f"Erreur pendant le téléchargement de {link} :(")




def get_song(i):
    clip = VideoFileClip(f'temp/{i}.{songs[i].extension}')

    if not check_sample(clip, songs[i]):
        raise ValueError(f'Sample too late for song {i} ({songs[i].info} from {songs[i].anime})')

    clip = clip.subclip(songs[i].sample - transition if i != len(songs)
                        else songs[i].sample,songs[i].sample + songs[i].sample_length + transition)
    layout = ImageClip(f'temp/a{i}.png').set_duration(clip.duration).set_position('center', 'center').resize(height=720)

    if clip.w / clip.h <= 16 / 9:
        return clip.resize(height=720), layout
    return clip.resize(width=1280), layout


def create_video(song_range, output_path):
    start, stop = song_range[0], song_range[1]
    video_layout = concatenate_videoclips(([images[start].crossfadein(transition)] +
                                           [images[i].crossfadeout(transition).crossfadein(transition)
                                            for i in range(start + 1, stop - 1)] +
                                           [images[stop - 1].crossfadeout(transition)])[::-1],
                                          padding=-transition,
                                          method='compose').set_position('center')
    video = concatenate_videoclips(([clips[start].crossfadein(transition).audio_fadein(transition)] +
                                    [clips[i].crossfadein(transition).audio_fadein(transition).crossfadeout(
                                        transition).audio_fadeout(transition)
                                     for i in range(start + 1, stop - 1)] +
                                    [clips[stop - 1].crossfadeout(transition).audio_fadeout(transition)])[::-1],
                                   padding=-transition,
                                   method='compose').set_position('center')
    video = video.subclip(0, video.duration)

    if clips[stop - 1].w / clips[stop - 1].h != 16 / 9:
        black = ImageClip('black.png').set_duration(clips[stop - 1].duration).set_position('center', 'center')
        pr_part = CompositeVideoClip([black, video, video_layout])
    else:
        pr_part = CompositeVideoClip([video, video_layout])
    pr_part.write_videofile(output_path, threads=4)

def check_sample(clip: VideoFileClip, song: SampledSong):
    if song is songs[1]:
        return song.sample + song.sample_length < clip.duration
    return song.sample + song.sample_length + transition < clip.duration


def normalize_audio(song_range):
    normalizer = FFmpegNormalize()
    for i in range(song_range[0], song_range[1]):
        normalizer.add_media_file(f'temp/{i}.wav', f'temp/normalized_{i}.wav')
    normalizer.run_normalization()
    for i in range(song_range[0], song_range[1]):
        clips[i].audio = AudioFileClip(f'temp/normalized_{i}.wav')


def split(x, n):
    r = x % n
    zp = n - r
    pp = x//n
    range_list = []
    start = 1
    for i in range(n):
        if i >= zp:
            range_list.append((start, start + pp + 1))
            start += pp + 1
        else:
            range_list.append((start, start + pp))
            start += pp
    return range_list


def get_progress(parts, range_list):
    try:
        with open('progress.json', 'r') as file:
            return json.load(file)
    except FileNotFoundError:
        return {
            'parts': parts,
            'range_list': range_list,
            'downloaded': 0,
            'done': 0,
            'layout': False,
            'video': False
        }


def save_progress(progress):
    with open(f"{pr_path}/résultats-nomi/progress.json", "w") as file:
        json.dump(progress, file)


def create_layouts(order, songs, output_path):
    people = len(order)
    C = nb_columns(order)
    pr_range = range(1, len(songs) + 1)

    import ScriptPRNomi as Nono
    Rang = list(pr_range)
    Total = [songs[i].score for i in pr_range]
    R = [[songs[i].scores[name] for name in order] for i in pr_range]
    Titre = [songs[i].anime + ' ' + songs[i].type for i in pr_range]
    Musique = [songs[i].info for i in pr_range]
    Picker = [songs[i].picker for i in pr_range]

    if people <= 8:
        Nono.creationimages8(R, C, Rang, Total, Titre, Musique, Picker, output_path, order)

    elif people <= 14:
        Nono.creationimages14(R, C, Rang, Total, Titre, Musique, Picker, output_path, order)

    elif people <= 18:
        Nono.creationimages18(R, C, Rang, Total, Titre, Musique, Picker, output_path, order)

    elif people <= 36:
        Nono.creationimages36(R, C, Rang, Total, Titre, Musique, Picker, output_path, order)        

    elif people <= 54:
        Nono.creationimages54(R, C, Rang, Total, Titre, Musique, Picker, output_path, order)


def fuse_parts(parts):
    video_parts = {}
    for i in range(parts):
        video_parts[i] = VideoFileClip(f'temp/part{i}.mp4')

    pr_video = concatenate_videoclips(([video_parts[0].crossfadein(transition).audio_fadein(transition)] +
                                       [video_parts[i].crossfadein(transition).audio_fadein(transition).crossfadeout(
                                        transition).audio_fadeout(transition)
                                     for i in range(1, parts - 1)] +
                                    [video_parts[parts - 1].crossfadeout(transition).audio_fadeout(transition)])[::-1],
                                   padding=-transition,
                                   method='compose').set_position('center')
    pr_video = pr_video.subclip(0, pr_video.duration)
    pr_video.write_videofile(f'{pr}.mp4', threads=4)


if __name__ == '__main__':
    scoring_pr = False
    base_path = os.getcwd()
    Path('temp').mkdir(parents=True, exist_ok=True)
    temp_path = os.path.join(base_path, 'temp')

    result_sheet = get_result_sheet()
    ws = result_sheet.active

    songs, order = get_results(ws)

    parts = len(songs) // song_per_part if not len(songs) % song_per_part else len(songs) // song_per_part + 1
    range_list = split(len(songs), parts)
    progress = get_progress(parts, range_list)

    pr_range = range(1, len(songs) + 1)

    if not progress['layout']:
        create_layouts(order, songs, temp_path)
        progress['layout'] = True
        save_progress(progress)

    while progress['done'] < progress['parts']:
        song_range = progress['range_list'][progress['done']]
        if progress['downloaded'] == progress['done']:
            os.chdir(temp_path)
            download_songs(songs, song_range)
            os.chdir(base_path)
            progress['downloaded'] += 1
            save_progress(progress)

        clips = {}
        images = {}

        for i in range(song_range[0], song_range[1]):
            (clips[i], images[i]) = get_song(i)
            audio = clips[i].audio
            audio.write_audiofile(f'temp/{i}.wav')

        normalize_audio(song_range)
        output_path = f'{pr}.mp4' if progress['parts'] == 1 else f'temp/part{progress["done"]}.mp4'
        create_video(song_range, output_path)
        progress['done'] += 1

        if progress['parts'] == 1:
            progress['video'] = True

        for i in range(song_range[0], song_range[1]):
            clips[i].close()
            images[i].close()

        save_progress(progress)

    if progress['done'] == progress['parts'] and not progress['video']:
        song_list = [songs[i] for i in pr_range]
        fuse_parts(progress['parts'])
        progress['video'] = True

        save_progress(progress)
