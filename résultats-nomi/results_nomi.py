from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.colors import Color
from pathlib import Path
import re
import pyperclip
from PIL import Image, ImageDraw, ImageFont
import process_PR_stats_ranking as affinity
from pyexcel_ods3 import save_data
from pyexcel_xlsx import get_data
import os
import configparser 

VERSION = "1.1.0"
print(f"results_nomi.py version {VERSION}")

print(f"Operating system: {os.name}")
regex = r'\\(.*) [(](.*)[)]' if os.name == 'nt' else r'/(.*) \((.*)\)'

# Données
pr = ''  # laisser vide si un seul pr dans le dossier
scoring_pr = False
order = []  # laisser vide si ordre alphabétique

config = configparser.ConfigParser()
config.read('../config.txt')
pr_path = config["general"]["pr_path"]
image_path = f"{pr_path}/pr-avatars"
results_path  = f"{pr_path}/résultats-nomi"

# Check de la configuration
config_error = False
for path in [pr_path, image_path, results_path]:
    if not os.path.isdir(path):
        config_error = True
        print(f"❌ Le dossier {path} n'existe pas, vérifiez votre configuration !")
if config_error:
    exit()

cadrenom = Image.open(f"{pr_path}/images/CadreNom.png")
cadrenom54 = Image.open(f"{pr_path}/images/CadreNom54.png")
layout = Image.open(f"{pr_path}/images/Layout.png")
layout54 = Image.open(f"{pr_path}/images/Layout54.png")
police_pseudo = f"{pr_path}/images/agencyfb.ttf"
save = f"{pr_path}/images/LayoutPR.png"
save_verify = f"{pr_path}/résultats-nomi/LayoutPR.png"
carre = Image.open(f"{pr_path}/images/carre.png")
police_pseudo2 = f"{pr_path}/images/MusticaPro-SemiBold 600.otf"

# Script

draw = ImageDraw.Draw(layout)
draw54 = ImageDraw.Draw(layout54)

avatars = []
fonta = ImageFont.truetype(police_pseudo, size=32)

class Song:
    def __init__(self, picker, anime, song_type, info, link):
        self.picker = picker
        self.anime = anime
        self.type = song_type
        self.info = info
        self.link = link
        self.scores = {}
        self.is_tied = False

    if scoring_pr:
        def __lt__(self, other):
            if self.score == other.score:
                self.is_tied, other.is_tied = True, True
            return self.score >= other.score
    else:
        def __lt__(self, other):
            if self.score == other.score:
                self.is_tied, other.is_tied = True, True
            return self.score <= other.score

    @property
    def score(self):
        if scoring_pr:
            return sum(100 * k for k in self.scores.values()) / 100
        return sum(self.scores.values())


def get_data_script(sheet, songs):
    ids = []
    for col in sheet.iter_cols(min_col=1, max_col=1, min_row=2):
        for cell in col:
            try:
                if cell.value:
                    ids.append(int(cell.value))
            except ValueError:
                break

    num_songs = max(ids)

    titres = sheet[1]
    titres = [titre.value for titre in titres]

    id_index = titres.index('ID')
    picker_index = titres.index('Picker')
    anime_index = titres.index('Anime Name')
    song_type_index = titres.index('Song Type')
    song_info_index = titres.index('Song Info')
    try :
        rank_index = titres.index('Rank') if not scoring_pr else titres.index('Score')
    except :
        exit("❌ Column Rank not found in the first sheet" if not scoring_pr else "❌ Column Score not found in the first sheet")

    for row in sheet.iter_rows(min_row=2, min_col=id_index, max_col=rank_index, max_row=num_songs + 1):
        if not row[anime_index].value:
            break
        if row[song_info_index].hyperlink:
            song = Song(row[picker_index].value ,row[anime_index].value, row[song_type_index].value, row[song_info_index].value,
                        row[song_info_index].hyperlink.target)
        else:
            song = Song(row[picker_index].value ,row[anime_index].value, row[song_type_index].value, row[song_info_index].value,
                        row[song_info_index].value)
        songs[row[id_index].value] = song

    return num_songs


def get_ranker_ranks(ws, num_songs, pseudo):
    looking_for = 'Rank' if not scoring_pr else 'Score'
    found = False
    ids = []
    ranks = []
    all_columns = {col[0].value for col in ws.columns}
    if not 'ID' in all_columns:
        raise Exception("Did not find ID column")
    if not 'Rank' in all_columns:
        raise Exception("Did not find Rank column")

    for col in ws.columns:
        if col[0].value is None:
            break
        if 'ID' == col[0].value:
            for row in list(col)[1:]:
                try:
                    ids.append(int(row.value))
                except:
                    break
        if looking_for == col[0].value:
            found = True
            for row in list(col)[1:]:
                try:
                    ranks.append(int(row.value)) if not scoring_pr else ranks.append(row.value)
                except:
                    if row.value is None:
                        pass
                    elif row.value.startswith("="):
                        print(f"⚠️ Animal {pseudo} put a formula in ranking cells")
                    pass

    ids = ids[:num_songs]
    ranks = ranks[:num_songs]


    if len(ranks) != num_songs:
        raise Exception(f"Did not score/rank all songs ({len(ranks)}/{num_songs})")
    if len(list(set(ranks))) != len(ranks) and not scoring_pr:
        raise Exception("One or more of the ranks is/are duplicated")
    if len(ids) != len(ranks):
        raise Exception("Ids, ranks column fill amount is not the same")
    if len(ids) != num_songs:
        raise Exception("Number of ids isn't the same as number of songs")
    if len(ranks) != num_songs:
        raise Exception("Number of ranks isn't the same as number of songs")
    if 0 in ranks and not scoring_pr:
        raise Exception("Rank 0 should not exist")
    if num_songs != max(ranks) and not scoring_pr:
        raise Exception("Max rank is higher than number of songs")

    return list(zip(ids, ranks, [pseudo] * num_songs))


def letter(col, col_abs=False):
    col_num = col
    col_num += 1
    col_str = ''
    col_abs = '$' if col_abs else ''

    while col_num:
        remainder = col_num % 26

        if remainder == 0:
            remainder = 26

        col_letter = chr(ord('A') + remainder - 1)
        col_str = col_letter + col_str
        col_num = int((col_num - 1) / 26)

    return col_abs + col_str


def pr_paths(path_list, pr):
    return [path for path in path_list if pr in str(path) and '(' in str(path)]


def make_order_from_json(people_list):
    for pseudo in people_list:
        order.append(pseudo)
        global C
        C = nb_columns(people_list)


def make_order(order, pr):
    if not len(order):
        path_list = pr_paths(Path('.').glob('**/*.xlsx'), pr)
        for path in path_list:
            pseudo = re.search(regex, str(path)).group(2)
            order.append(pseudo)
            
        order.sort()
            
            
def verify_pfp(order):
    accepted = True
    for pseudo in order:
        if not os.path.exists(f'{image_path}/{pseudo}.png'):
            accepted = False
            print(f"❌ Avatar for {pseudo} not found")
    return accepted


def fill_values(sheet, order, song_list, final_sheet, scoring_pr):
    row1 = ['Rank', 'Picker', 'Anime Name', 'Song Type', 'Song Info', 'Score',]
    if scoring_pr:
        row1.append('Average')
    if not final_sheet:
        row1.append('Sample (seconds)')
        row1.append('Sample length (seconds)')

    part1 = len(row1)
    for i in range(part1):
        header = sheet.cell(1, i + 1, row1[i])
        header.font = Font(bold=True)
        header.fill = gray_background

    for i in range(len(order)):
        name = sheet.cell(1, i + part1 + 1, order[i])
        name.fill = gray_background

    for i in range(len(song_list)):
        if not song_list[i].is_tied or final_sheet:
            sheet.cell(i + 2, 1, i + 1)
        sheet.cell(i + 2, 2, song_list[i].picker)
        sheet.cell(i + 2, 3, song_list[i].anime)
        sheet.cell(i + 2, 4, song_list[i].type)

        link_cell = sheet.cell(i + 2, 5, song_list[i].info)
        link_cell.hyperlink = song_list[i].link
        link_cell.font = Font(color=link_color, underline='single')

        sheet.cell(i + 2, 6, song_list[i].score)
        if scoring_pr:
            sheet.cell(i + 2, 7, float('{:.2f}'.format(song_list[i].score / len(order))))

        for j in range(len(order)):
            sheet.cell(i + 2, j + part1 + 1, song_list[i].scores[order[j]])
            
def resize_columns(sheet, final_sheet=False):
    dims = {}
    for row in sheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        if final_sheet:
            if col in 'ABCDE' or scoring_pr and col == 'F':
                sheet.column_dimensions[col].width = value + 5
        else:
            if col in 'ABCDEFG' or scoring_pr and col == 'H':
                sheet.column_dimensions[col].width = value + 5


def color_sheet(sheet, song_list, order, scoring_pr):
    i = 1
    for row in sheet.rows:
        if i % 2 and i > 1:
            for cell in row:
                cell.fill = light_gray_background
        elif not i % 2:
            for cell in row:
                cell.fill = white_background
        i += 1

    for i in range(len(song_list)):
        sheet.conditional_formatting.add(f'F{i + 2}:{letter(4 + len(order))}{i + 2}', CellIsRule(operator='=',
                                                                                                 formula=[
                                                                                                     f'=MIN($F${i + 2}:${letter(4 + len(order))}${i + 2})'],
                                                                                                 fill=green_background if not scoring_pr else red_background))
        sheet.conditional_formatting.add(f'F{i + 2}:{letter(4 + len(order))}{i + 2}', CellIsRule(operator='=',
                                                                                                 formula=[
                                                                                                     f'=MAX($F${i + 2}:${letter(4 + len(order))}${i + 2})'],
                                                                                                 fill=red_background if not scoring_pr else green_background))


def create_results_sheet(pr, order, song_list, scoring_pr, make_sheet, outputpath, final_sheet=False):
    os.chdir(outputpath)
    if not make_sheet:
        pyperclip.copy(to_clipboard + '```')
        exit('❌ Stallers, could not finish processing')

    result = Workbook()
    sheet = result.active

    fill_values(sheet, order, song_list, final_sheet, scoring_pr)
    resize_columns(sheet, final_sheet)

    if final_sheet:
        color_sheet(sheet, song_list, order, scoring_pr)
    sheet.auto_filter.ref = sheet.dimensions
    result_sheet_name = f'{pr} results.xlsx' if final_sheet else f'{pr} samples.xlsx'
    result.save(result_sheet_name)
    

def fill_avatars(C, order, size):
    i = 0
    for j in range(len(C)):
        provisoire = []
        while len(provisoire) != C[j]:
            pp = Image.open(f'{image_path}/{order[i]}.png')
            pp = pp.resize((size, size))
            provisoire.append(pp)
            i += 1
        avatars.append(provisoire)
        
    return avatars

    
def worryheart(people):
    # On convertit les jpg en png
    jpg_files = [file for file in os.listdir(image_path) if file.lower().endswith('.jpg') or file.lower().endswith('.jpeg')]
    for jpg_file in jpg_files:
        exit_code = os.system(f'ffmpeg -y -i "{image_path}/{jpg_file}" "{image_path}/{jpg_file.split(".")[0]}.png" > NUL 2>&1')
        if exit_code != 0:
            print(f"Command failed with exit code {exit_code}.")

    global LINE1
    if people <= 8:
        layout.paste(cadrenom, mask = cadrenom)
        fontb = ImageFont.truetype(police_pseudo2, size=35)
        avatars = fill_avatars(C, order, 152)

        incr = 111
        for i in range(C[0]):
            layout.paste(avatars[0][i], (10,incr))
            draw.text((86,incr-3),order[i],fill='white', stroke_fill='black', stroke_width=1, font=fontb,anchor="ms")
            incr += 91+152
        
        incr = 111
        for i in range(C[1]):
            layout.paste(avatars[1][i], (1663,incr))
            draw.text((1747,incr-3),order[i+C[0]],fill='white', stroke_fill='black', stroke_width=1, font=fontb,anchor="ms")
            incr += 91+152

    elif people <= 14:
        layout.paste(cadrenom, mask = cadrenom)
        avatars = fill_avatars(C, order, 118)

        incr = 49
        for i in range(C[0]):
            layout.paste(avatars[0][i], (37,incr))
            draw.text((96,incr+3),order[i],fill='white', stroke_fill='black', stroke_width=1, font=fonta,anchor="ms")
            incr += 145
        
        incr = 49
        for i in range(C[1]):
            layout.paste(avatars[1][i], (1690,incr))
            draw.text((1748,incr+3),order[i+C[0]],fill='white', stroke_fill='black', stroke_width=1, font=fonta,anchor="ms")
            incr += 145
 
    elif 15 <= people <= 18:
        layout.paste(cadrenom, mask = cadrenom)
        avatars = fill_avatars(C, order, 97)

        incr = 22
        for i in range(C[0]):
            layout.paste(avatars[0][i], (19,incr))
            draw.text((68,incr+3),order[i],fill='white', stroke_fill='black', stroke_width=1, font=fonta,anchor="ms")
            incr += 23+97
        
        incr = 22
        for i in range(C[1]):
            layout.paste(avatars[1][i], (1673,incr))
            draw.text((1722,incr+3),order[i+C[0]],fill='white', stroke_fill='black', stroke_width=1, font=fonta,anchor="ms")
            incr += 23+97
 
    elif 19 <= people <= 36:
        layout.paste(cadrenom, mask = cadrenom)
        avatars = fill_avatars(C, order, 97)

        incr = 22
        incrcar = 70
        for i in range(C[0]):
            layout.paste(avatars[0][i], (19,incr))
            layout.paste(carre, (0,incrcar),mask=carre)
            draw.text((68,incr+3),order[i],fill='white', stroke_fill='black', stroke_width=1, font=fonta,anchor="ms")
            incr += 23+97
            incrcar += 120
    
        incr = 22
        incrcar = 70
        for i in range(C[1]):
            layout.paste(avatars[1][i], (149,incr))
            layout.paste(carre, (130,incrcar),mask=carre)
            draw.text((198,incr+3),order[i+C[0]],fill='white', stroke_fill='black', stroke_width=1, font=fonta,anchor="ms")
            incr += 23+97
            incrcar += 120
    
        incr = 22
        incrcar = 70
        for i in range(C[2]):
            layout.paste(avatars[2][i], (1673,incr))
            layout.paste(carre, (1655,incrcar),mask=carre)
            draw.text((1722,incr+3),order[i+C[0]+C[1]],fill='white', stroke_fill='black', stroke_width=1, font=fonta,anchor="ms")
            incr += 23+97
            incrcar += 120

        incr = 22
        incrcar = 70
        for i in range(C[3]):
            layout.paste(avatars[3][i], (1803,incr))
            layout.paste(carre, (1785,incrcar),mask=carre)  
            draw.text((1852,incr+3),order[i+C[0]+C[1]+C[2]],fill='white', stroke_fill='black', stroke_width=1, font=fonta,anchor="ms")
            incr += 23+97
            incrcar += 120
            
    elif 37 <= people <= 54:
        layout54.paste(cadrenom54, mask = cadrenom54)
        avatars = fill_avatars(C, order, 97)

        incr = 22
        incrcar = 70
        for i in range(C[0]):
            layout54.paste(avatars[0][i], (19,incr))
            layout54.paste(carre, (0,incrcar),mask=carre)
            draw54.text((68,incr+3),order[i],fill='white', stroke_fill='black', stroke_width=1, font=fonta,anchor="ms")
            incr += 23+97
            incrcar += 120
    
        incr = 22
        incrcar = 70
        for i in range(C[1]):
            layout54.paste(avatars[1][i], (149,incr))
            layout54.paste(carre, (130,incrcar),mask=carre)
            draw54.text((198,incr+3),order[i+C[0]],fill='white', stroke_fill='black', stroke_width=1, font=fonta,anchor="ms")
            incr += 23+97
            incrcar += 120
    
        incr = 22
        incrcar = 70
        for i in range(C[2]):
            layout54.paste(avatars[2][i], (279,incr))
            layout54.paste(carre, (260,incrcar),mask=carre)
            draw54.text((328,incr+3),order[i+C[0]+C[1]],fill='white', stroke_fill='black', stroke_width=1, font=fonta,anchor="ms")
            incr += 23+97
            incrcar += 120
    
        incr = 22
        incrcar = 70
        for i in range(C[3]):
            layout54.paste(avatars[3][i], (1543,incr))
            layout54.paste(carre, (1525,incrcar),mask=carre)
            draw54.text((1592,incr+3),order[i+C[0]+C[1]+C[2]],fill='white', stroke_fill='black', stroke_width=1, font=fonta,anchor="ms")
            incr += 23+97
            incrcar += 120

        incr = 22
        incrcar = 70
        for i in range(C[4]):
            layout54.paste(avatars[4][i], (1673,incr))
            layout54.paste(carre, (1655,incrcar),mask=carre)
            draw54.text((1722,incr+3),order[i+C[0]+C[1]+C[2]+C[3]],fill='white', stroke_fill='black', stroke_width=1, font=fonta,anchor="ms")
            incr += 23+97
            incrcar += 120

        incr = 22
        incrcar = 70
        for i in range(C[5]):
            layout54.paste(avatars[5][i], (1803,incr))
            layout54.paste(carre, (1785,incrcar),mask=carre)
            draw54.text((1852,incr+3),order[i+C[0]+C[1]+C[2]+C[3]+C[4]],fill='white', stroke_fill='black', stroke_width=1, font=fonta,anchor="ms")
            incr += 23+97
            incrcar += 120

        
    if people <= 36:
        layout.save(save)
        layout.save(save_verify)
    elif 36 <= people <=54:
        layout54.save(save)
        layout54.save(save_verify)

    else:
        exit("❌ Trop de participants")
            
def nb_columns(order):
    C = []
    people = len(order)
    if people <= 18:
        if people % 2 == 0:
            for i in range(2):
                C.append(int(people / 2))
            return C
        elif people % 2 == 1:
            people -= 1
            C.append(int(people / 2 + 1))
            C.append(int(people / 2))
            return C
    elif people <= 36:
        if people % 4 == 0:
            for i in range(4):
                C.append(int(people / 4))
            return C
        elif people % 4 == 1:
            people -= 1
            C.append(int(people / 4 + 1))
            for i in range(3):
                C.append(int(people / 4))
            return C
        elif people % 4 == 2:
            people -= 2
            for i in range(2):
                C.append(int(people / 4 + 1))
            for i in range(2):
                C.append(int(people / 4))
            return C
        elif people % 4 == 3:
            people -= 3
            for i in range(3):
                C.append(int(people / 4 + 1))
            C.append(int(people / 4))
            return C
    else:
        if people % 6 == 0:
            for i in range(6):
                C.append(int(people / 6))
            return C
        elif people % 6 == 1:
            people -= 1
            C.append(int(people / 6 + 1))
            for i in range(5):
                C.append(int(people / 6))
            return C
        elif people % 6 == 2:
            people -= 2
            for i in range(2):
                C.append(int(people / 6 + 1))
            for i in range(4):
                C.append(int(people / 6))
            return C
        elif people % 6 == 3:
            people -= 3
            for i in range(3):
                C.append(int(people / 6 + 1))
            for i in range(3):
                C.append(int(people / 6))
            return C
        elif people % 6 == 4:
            people -= 4
            for i in range(4):
                C.append(int(people / 6 + 1))
            for i in range(2):
                C.append(int(people / 6))
            return C
        elif people % 6 == 5:
            people -= 5
            for i in range(5):
                C.append(int(people / 6 + 1))
            C.append(int(people / 6))
            return C


def pr_find():
    paths = Path('.').glob('**/*.xlsx')
    for base_file in paths:
        if '(' in str(base_file):
            pr_name = re.search(regex, str(base_file)).group(1)
            break
    if "pr_name" not in locals():
        exit("❌ Pas de dossier de PR trouvé :(")
    return pr_name

def get_affinity(outputpath):
    os.chdir(outputpath)
    data_xlsx = get_data(f'{pr} results.xlsx')
    save_data(f'{pr}.ods', data_xlsx)
    affinity.main()
    os.remove(f'{pr}.ods')

white = 'FFFFFF'
light_gray = 'F3F3F3'
link_color = "1155cc"

cell_background_color = "cccccc"
light_gray_background = PatternFill(patternType='solid', fgColor=Color(rgb=light_gray))
white_background = PatternFill(patternType='solid', fgColor=Color(rgb=white))
gray_color = Color(rgb=cell_background_color)
gray_background = PatternFill(patternType="solid", fgColor=gray_color)

green = '00FF00'
green_background = PatternFill(patternType='solid', bgColor=Color(rgb=green))
red = 'FF0000'
red_background = PatternFill(patternType='solid', bgColor=Color(rgb=red))



if __name__ == '__main__':
    path_list = pr_paths(Path('.').glob('**/*.xlsx'), pr)
    first = True
    make_sheet = True
    songs = {}
    to_clipboard = '```\n'
    LINE1 = []

    if not pr:
        pr = pr_find()

    make_order(order, pr)
    
    if not verify_pfp(order):
        exit("❌ Missing avatars")
        
    C = nb_columns(order)
    for path in path_list:
        wb = load_workbook(path)
        ws = wb.active
        sheet_name = str(path)
        pseudo = re.search(regex, sheet_name).group(2)

        if first:
            num_songs = get_data_script(ws, songs)
            first = False

        try:
            rank_list = get_ranker_ranks(ws, num_songs, pseudo)
            for song_id, rank, name in rank_list:
                songs[song_id].scores[name] = rank
        except Exception as e:
            message = f'⚠️ {pseudo} : {e}'
            print(message)
            to_clipboard += message + '\n'
            make_sheet = False
            
    song_list = list(songs.values())
    song_list.sort()

    create_results_sheet(pr, order, song_list, scoring_pr, make_sheet, results_path)
    print(len(order))
    print(order)
    worryheart(len(order))
    
    try:
        base_path = os.getcwd()
        Path(f'{pr} Stats et Sheet').mkdir(parents=True, exist_ok=True)
        temp_path = os.path.join(base_path, f'{pr} Stats et Sheet')
        os.chdir(temp_path)
        create_results_sheet(pr, order, song_list, scoring_pr, make_sheet, temp_path, final_sheet=True)
        get_affinity(temp_path)
        print("Affinité Faite")
    except Exception as e:
        print(e)
        pass


