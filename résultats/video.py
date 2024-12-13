from results import nb_columns, create_results_sheet, worryheart, make_order_from_json, Song
from openpyxl import load_workbook
from pathlib import Path
from moviepy.editor import *
import os
from ffmpeg_normalize import FFmpegNormalize
import json
import argparse
import configparser
import requests
import aiohttp
import asyncio
from yt_dlp import YoutubeDL


VERSION = "1.1.0"
print(f"video.py version {VERSION}")

transition = 1
pr = ''
song_per_part = 45
reverse = False

config = configparser.ConfigParser()
config.read('../config.txt')
pr_path = config["general"]["pr_path"]
image_path = f'{pr_path}/pr-avatars'
results_path  = f"{pr_path}/résultats"

# Check de la configuration
config_error = False
for path in [pr_path, image_path, results_path]:
    if not os.path.isdir(path):
        config_error = True
        print(f"❌ Le dossier {path} n'existe pas, vérifiez votre configuration !")
if config_error:
    exit()


class SampledSong(Song):
    def __init__(self, anime, song_type, info, link, score, sample, sample_length, ranks, order):
        Song.__init__(self, anime, song_type, info, link)
        self.final_score = score
        if sample is None:
            exit("❌ Une ou plusieurs des cases Sample est vide!")
        self.sample = max(sample, transition)
        self.sample_length = sample_length
        if sample_length is None:
            exit("❌ Une ou plusieurs des cases Sample Length est vide!")
        if sample_length <= 0:
            exit("❌ Une ou plusieurs des cases Sample Length contient une valeur invalide ! (<= 0)")
        self.scores = {order[i]: ranks[i] for i in range(len(order))}
        if 'catbox' in link or 'animemusicquiz.com' in link:
            self.extension = get_extension(link)
        elif 'youtu' in link:
            self.extension = 'mp4'


def get_result_sheet():
    global pr
    paths = Path('.').glob('**/*.xlsx')
    for path in paths:
        if 'samples' in str(path):
            pr = str(path).split()[0]
            return load_workbook(path)


def get_results(sheet):
    global scoring_pr
    rows = []
    songs = {}
    for row in sheet.iter_rows():
        to_add = []
        for cell in row:
            to_add.append(cell.value)
            if cell.hyperlink:
                to_add.append(cell.hyperlink.target)
        rows.append(to_add)

    header = rows[0]
    scoring_pr = 'Average' in header

    rank_index = header.index('Rank')
    anime_index = header.index('Anime Name')
    type_index = header.index('Song Type')
    info_index = header.index('Song Info')
    score_index = header.index('Score') + 1
    sample_index = header.index('Sample (seconds)') + 1
    sample_length_index = header.index('Sample length (seconds)') + 1
    people_index = max(rank_index, anime_index, type_index, info_index,
                       score_index, sample_index, sample_length_index) + 1

    order = rows[0][people_index-1:]

    for i in range(1, len(rows)):
        if not rows[i][rank_index]:
            raise Exception('Tiebreak needed')
        songs[rows[i][rank_index]] = SampledSong(rows[i][anime_index], rows[i][type_index], rows[i][info_index],
                                                 rows[i][info_index + 1], rows[i][score_index], rows[i][sample_index],
                                                 rows[i][sample_length_index], rows[i][people_index:], order)
    return songs, order


def get_extension(link):
    return link.split(".")[-1]


def youtube_dl(link, output_name):
    ydl_opts = {
        'format': 'bestvideo[ext=mp4]+bestaudio[ext=m4a]/best[ext=mp4]/best',
        'outtmpl': f'{output_name}',
        'merge_output_format': 'mp4',
        'postprocessors': [{
            'key': 'FFmpegVideoConvertor',
            'preferedformat': 'mp4'
        }]
    }

    with YoutubeDL(ydl_opts) as ydl:
        ydl.download([link])

async def download_file(session, url, filename):
    async with session.get(url) as response:
        if response.status == 200:
            with open(filename, 'wb') as file:
                file.write(await response.read())
        else:
            print(f"Failed to download {url}")

async def download_songs_async(songs, song_range):
    async with aiohttp.ClientSession() as session:
        tasks = []
        for i in range(song_range[0], song_range[1]):
            link = songs[i].link
            print("Downloading: " + link)
            if 'catbox' in link or 'animemusicquiz.com' in link:
                tasks.append(download_file(session, link, f'{i}.{songs[i].extension}'))
            elif 'youtu' in link:
                tasks.append(asyncio.to_thread(youtube_dl, link, f'{i}.mp4'))
        await asyncio.gather(*tasks)

def download_songs(songs, song_range):
    asyncio.run(download_songs_async(songs, song_range))


def get_song(i):
    clip = VideoFileClip(f'temp/{i}.{songs[i].extension}')

    if not check_sample(clip, songs[i]):
        raise ValueError(f'Sample too late for song {i} ({songs[i].info} from {songs[i].anime})')

    clip = clip.subclip(songs[i].sample - transition if i != len(songs)
                        else songs[i].sample,songs[i].sample + songs[i].sample_length + transition)
    layout = ImageClip(f'temp/a{i}.png').set_duration(clip.duration).set_position('center', 'center').resize(height=720)

    if clip.w / clip.h <= 16 / 9:
        return clip.resize(height=720), layout
    return clip.resize(width=1280), layout


def create_video(song_range, output_path):
    start, stop = song_range[0], song_range[1]
    video_layout = concatenate_videoclips(([images[start].crossfadein(transition)] +
                                           [images[i].crossfadeout(transition).crossfadein(transition)
                                            for i in range(start + 1, stop - 1)] +
                                           [images[stop - 1].crossfadeout(transition)])[::-1],
                                          padding=-transition,
                                          method='compose').set_position('center')
    video = concatenate_videoclips(([clips[start].crossfadein(transition).audio_fadein(transition)] +
                                    [clips[i].crossfadein(transition).audio_fadein(transition).crossfadeout(
                                        transition).audio_fadeout(transition)
                                     for i in range(start + 1, stop - 1)] +
                                    [clips[stop - 1].crossfadeout(transition).audio_fadeout(transition)])[::-1],
                                   padding=-transition,
                                   method='compose').set_position('center')
    video = video.subclip(0, video.duration)

    if clips[stop - 1].w / clips[stop - 1].h != 16 / 9:
        black = ImageClip('black.png').set_duration(clips[stop - 1].duration).set_position('center', 'center')
        pr_part = CompositeVideoClip([black, video, video_layout])
    else:
        pr_part = CompositeVideoClip([video, video_layout]) 
    pr_part.write_videofile(output_path, threads=4)

def check_sample(clip: VideoFileClip, song: SampledSong):
    if song is songs[1]:
        return song.sample + song.sample_length < clip.duration
    return song.sample + song.sample_length + transition < clip.duration


def normalize_audio(song_range):
    normalizer = FFmpegNormalize()
    for i in range(song_range[0], song_range[1]):
        normalizer.add_media_file(f'temp/{i}.wav', f'temp/normalized_{i}.wav')
    normalizer.run_normalization()
    for i in range(song_range[0], song_range[1]):
        clips[i].audio = AudioFileClip(f'temp/normalized_{i}.wav')


def split(x, n):
    r = x % n
    zp = n - r
    pp = x//n
    range_list = []
    start = 1
    for i in range(n):
        if i >= zp:
            range_list.append((start, start + pp + 1))
            start += pp + 1
        else:
            range_list.append((start, start + pp))
            start += pp
    return range_list


def get_progress(parts, range_list):
    try:
        with open('progress.json', 'r') as file:
            return json.load(file)
    except FileNotFoundError:
        return {
            'parts': parts,
            'range_list': range_list,
            'downloaded': 0,
            'done': 0,
            'layout': False,
            'video': False
        }


def save_progress(progress):
    with open('progress.json', 'w') as file:
        json.dump(progress, file, indent=4)


def create_layouts(order, songs, output_path):
    people = len(order)
    C = nb_columns(len(order))
    pr_range = range(1, len(songs) + 1)

    import ScriptPR as Nono
    Rang = list(pr_range)
    Total = [songs[i].score for i in pr_range]
    R = [[songs[i].scores[name] for name in order] for i in pr_range]
    Titre = [songs[i].anime + ' ' + songs[i].type for i in pr_range]
    Musique = [songs[i].info for i in pr_range]

    if people ==1:
        Nono.creationimagessolo(R, C, Rang, Total, Titre, Musique, output_path)
    elif people <= 8:
        Nono.creationimages8(R, C, Rang, Total, Titre, Musique, output_path)

    elif people <= 14:
        Nono.creationimages14(R, C, Rang, Total, Titre, Musique, output_path)

    elif people <= 18:
        Nono.creationimages18(R, C, Rang, Total, Titre, Musique, output_path)

    elif people <= 36:
        Nono.creationimages36(R, C, Rang, Total, Titre, Musique, output_path)        

    elif people <= 54:
        Nono.creationimages54(R, C, Rang, Total, Titre, Musique, output_path)

    elif people <= 60:
        Nono.creationimages60(R, C, Rang, Total, Titre, Musique, output_path)

    elif people <=72:
        Nono.creationimages72(R, C, Rang, Total, Titre, Musique, output_path)

    elif people <= 80:
        Nono.creationimages80(R, C, Rang, Total, Titre, Musique, output_path)

def fuse_parts(parts):
    video_parts = {}
    for i in range(parts):
        video_parts[i] = VideoFileClip(f'temp/part{i}.mp4')

    pr_video = concatenate_videoclips(([video_parts[0].crossfadein(transition).audio_fadein(transition)] +
                                       [video_parts[i].crossfadein(transition).audio_fadein(transition).crossfadeout(
                                        transition).audio_fadeout(transition)
                                     for i in range(1, parts - 1)] +
                                    [video_parts[parts - 1].crossfadeout(transition).audio_fadeout(transition)])[::-1],
                                   padding=-transition,
                                   method='compose').set_position('center')
    pr_video = pr_video.subclip(0, pr_video.duration)
    pr_video.write_videofile(f'{pr}.mp4', threads=4)
    os.remove("progress.json")


def download_json_profile_pic(user):
    print(f"Téléchargement de la PP de {user['name']}...")
    response = requests.get(user['image'])
    if response.status_code == 200:
        with open(f"{image_path}/{user['name']}.png", 'wb') as file:
            file.write(response.content)
        print(f"PP de {user['name']} téléchargée !")
    else:
        print(f"Failed to retrieve PP. Status code: {response.status_code}")

def process_json(path):
    with open(path, 'r', encoding='utf-8') as json_file:
        data = json.load(json_file)
        songs = {}
        last_order = []
        global pr
        pr = data["name"]
        accepted = True
        for user in data["voters"]:
            download_json_profile_pic(user)
            if not os.path.exists(f"{image_path}/{user['name']}.png"):
                print(f"❌ PP de {user['name']} non trouvée !")
                accepted = False
        if not accepted:
            exit("❌ Une ou plusieurs PP n'ont pas été trouvées !")
            
        data["songList"] = sorted(data["songList"], key=lambda x: x["rankPosition"], reverse=reverse)
        for i in range(0, len(data["songList"])):
            current_song = data["songList"][i]
            ranks = [vote["rank"] for vote in current_song["voters"]]
            order = [vote["name"] for vote in current_song["voters"]]
            last_order = order
            songs[i+1] = SampledSong(current_song['source'] if current_song['source'] != '' and current_song['source'] is not None else '', f"- {current_song['type']}" if current_song['source'] != '' and current_song['source'] is not None else f"{current_song['type']}", current_song["artist"] + " - " + current_song["title"], current_song["urlVideo"], current_song["totalRank"], current_song["startSample"], current_song["sampleLength"], ranks, order)
    return songs, last_order

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--json', type=str, required=False, help="JSON file path")
    parser.add_argument('--song_per_part', type=int, required=False, help="Number of songs per part")
    parser.add_argument('--reverse', help="Reverse order for rank (only available for JSON)", action='store_true')
    args = parser.parse_args()

    scoring_pr = False
    base_path = os.getcwd()
    Path('temp').mkdir(parents=True, exist_ok=True)
    temp_path = os.path.join(base_path, 'temp')
    
    if args.reverse:
        reverse = True
    if args.song_per_part:
        song_per_part = args.song_per_part

    if args.json is not None:
        songs, order = process_json(args.json)
        make_order_from_json(order)
        worryheart(len(order))
    else :
        result_sheet = get_result_sheet()
        ws = result_sheet.active

        songs, order = get_results(ws)
    parts = len(songs) // song_per_part if not len(songs) % song_per_part else len(songs) // song_per_part + 1
    range_list = split(len(songs), parts)
    progress = get_progress(parts, range_list)
    pr_range = range(1, len(songs) + 1)

    if not progress['layout']:
        create_layouts(order, songs, temp_path)
        print("Les layouts ont été créés")
        progress['layout'] = True
        save_progress(progress)

    while progress['done'] < progress['parts']:
        song_range = progress['range_list'][progress['done']]
        if progress['downloaded'] == progress['done']:
            os.chdir(temp_path)
            download_songs(songs, song_range)
            os.chdir(base_path)
            progress['downloaded'] += 1
            save_progress(progress)

        clips = {}
        images = {}

        for i in range(song_range[0], song_range[1]):
            (clips[i], images[i]) = get_song(i)
            audio = clips[i].audio
            audio.write_audiofile(f'temp/{i}.wav')

        normalize_audio(song_range)
        output_path = f'{pr}.mp4' if progress['parts'] == 1 else f'temp/part{progress["done"]}.mp4'

        create_video(song_range, output_path)
        progress['done'] += 1

        if progress['parts'] == 1:
            progress['video'] = True

        for i in range(song_range[0], song_range[1]):
            clips[i].close()
            images[i].close()

        save_progress(progress)

    if progress['done'] == progress['parts'] and not progress['video']:
        #song_list = [songs[i] for i in pr_range]
            
        #get_affinity()

        fuse_parts(progress['parts'])
        progress['video'] = True

        save_progress(progress)
