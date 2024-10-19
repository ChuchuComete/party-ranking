import argparse
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter


def convert_json_to_excel(args):
    with open(args.json, 'r') as f:
        data = json.load(f)
        
    # Sort by rankPosition descending
    data['songList'] = sorted(data['songList'], key=lambda x: x['rankPosition'], reverse=True)
    
    # Extract relevant data from JSON
    song_list = data['songList']
    
    # Create a DataFrame from the JSON data
    rows = []
    for song in song_list:
        row = {
            'Rank': song['rankPosition'],
            'Anime': song['anime'],
            'Song Type': song['type'],
            'Song': song['artist'] + ' - ' + song['title'],
            'Score': song['totalRank']
        }
        for voter in song['voters']:
            row[voter['name']] = voter['rank']
        rows.append(row)
    
    dataframe = pd.DataFrame(rows)
    
    print("\nScore grid based on the settings you entered, check if it's correct!\n")
    print(dataframe)
    
    # Save DataFrame to Excel with openpyxl
    with pd.ExcelWriter('results.xlsx', engine='openpyxl') as writer:
        dataframe.to_excel(writer, index=False, sheet_name='Results')
        worksheet = writer.sheets['Results']
        
        # Add a table to the worksheet
        tab = Table(displayName="Table1", ref=worksheet.dimensions)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        worksheet.add_table(tab)
        
        # Add autofilter to the worksheet
        worksheet.auto_filter.ref = worksheet.dimensions
        
        # Define the fill colors for the best and worst ranks
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        
        # Apply conditional formatting for the best and worst ranks
        for col in worksheet.iter_cols(min_col=3, max_col=worksheet.max_column, min_row=2, max_row=worksheet.max_row):
            col_letter = col[0].column_letter
            worksheet.conditional_formatting.add(f'{col_letter}2:{col_letter}{worksheet.max_row}',
                                                 CellIsRule(operator='equal', formula=['MIN(${}$2:${}${})'.format(col_letter, col_letter, worksheet.max_row)], fill=green_fill))
            worksheet.conditional_formatting.add(f'{col_letter}2:{col_letter}{worksheet.max_row}',
                                                 CellIsRule(operator='equal', formula=['MAX(${}$2:${}${})'.format(col_letter, col_letter, worksheet.max_row)], fill=red_fill))
            
        # Add hyperlinks to the 'Song' column
        for row in range(2, worksheet.max_row + 1):
            song_cell = worksheet[f'D{row}']
            song_cell.hyperlink = song_list[row - 2]['urlVideo']
            song_cell.style = "Hyperlink"
    
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--json', type=str, required=True, help="JSON file path")
    args = parser.parse_args()
    
    convert_json_to_excel(args)
    
if __name__ == '__main__':
    main()
    